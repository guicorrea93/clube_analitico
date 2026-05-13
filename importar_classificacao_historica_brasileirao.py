"""Importa classificacoes finais historicas para o SQLite.

Entrada principal: classificacao_codex.xlsx

O script e idempotente: recria apenas as tabelas de staging/fato historica
e preserva as tabelas de partidas/classificacao rodada-a-rodada existentes.
"""
from __future__ import annotations

import re
import shutil
import sqlite3
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).parent
DB = ROOT / "db" / "brasileirao.db"
XLSX = ROOT / "classificacao_codex.xlsx"
REPORT = ROOT / "relatorio_importacao_classificacao_final.xlsx"
BACKUP = ROOT / "db" / "brasileirao_before_classificacao_final.db"


def key(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def canonical_seed() -> dict[str, str]:
    aliases = {
        "border link ceara alt ceara": "Ceara",
        "abc": "ABC",
        "asa": "ASA",
        "alecrim": "Alecrim",
        "alianca": "Aliança",
        "america mg": "America-MG",
        "america mineiro": "America-MG",
        "america rj": "America-RJ",
        "america gb": "America-RJ",
        "america rn": "America-RN",
        "america ce": "América-CE",
        "america se": "América-SE",
        "america sp": "América-SP",
        "americano": "Americano-RJ",
        "americano rj": "Americano-RJ",
        "anapolina": "Anapolina",
        "anapolis": "Anápolis",
        "athletico paranaense": "Athletico-PR",
        "athletico pr": "Athletico-PR",
        "atletico paranaense": "Athletico-PR",
        "atletico pr": "Athletico-PR",
        "atletico go": "Atletico-GO",
        "atletico mineiro": "Atletico-MG",
        "atletico mg": "Atletico-MG",
        "auto esporte": "Auto Esporte",
        "auto esporte pi": "Auto Esporte-PI",
        "avai": "Avai",
        "bahia": "Bahia",
        "bahia ba": "Bahia",
        "bangu": "Bangu",
        "barueri": "Barueri",
        "botafogo": "Botafogo-RJ",
        "botafogo rj": "Botafogo-RJ",
        "botafogo pb": "Botafogo-PB",
        "botafogo sp": "Botafogo-SP",
        "bragantino": "Bragantino",
        "red bull bragantino": "Bragantino",
        "red bull bragantino sp": "Bragantino",
        "brasil de pelotas": "Brasil de Pelotas",
        "brasil pelotas": "Brasil de Pelotas",
        "brasiliense": "Brasiliense",
        "brasilia": "Brasília EC",
        "brasilia ec": "Brasília EC",
        "ceub": "CEUB",
        "crb": "CRB",
        "csa": "CSA",
        "caldense": "Caldense",
        "campinense": "Campinense",
        "campo grande": "Campo Grande-RJ",
        "campo grande rj": "Campo Grande-RJ",
        "capelense al": "Capelense-AL",
        "catuense": "Catuense",
        "caxias": "Caxias",
        "ceara": "Ceara",
        "central": "Central",
        "chapecoense": "Chapecoense",
        "chapecoense sc": "Chapecoense",
        "colatina": "Colatina",
        "colorado": "Colorado-PR",
        "colorado pr": "Colorado-PR",
        "comercial": "Comercial-SP",
        "comercial ms": "Comercial-MS",
        "comercial mt": "Comercial-MS",
        "comercial pr": "Comercial-PR",
        "comercial rp": "Comercial-RP",
        "comercial sp": "Comercial-RP",
        "confianca": "Confiança",
        "corinthians": "Corinthians",
        "corinthians sp": "Corinthians",
        "coritiba": "Coritiba",
        "coritiba pr": "Coritiba",
        "corumbaense": "Corumbaense",
        "criciuma": "Criciuma",
        "cruzeiro": "Cruzeiro",
        "cruzeiro mg": "Cruzeiro",
        "cruzeiro do sul df": "Cruzeiro do Sul-DF",
        "cuiaba": "Cuiaba",
        "defele df": "Defelê-DF",
        "deportiva es": "Desportiva-ES",
        "desportiva": "Desportiva-ES",
        "desportiva es": "Desportiva-ES",
        "dom bosco": "Dom Bosco-MT",
        "dom bosco mt": "Dom Bosco-MT",
        "eletrovapo rj": "Eletrovapo-RJ",
        "estrela do mar pb": "Estrela do Mar-PB",
        "fast": "Fast-AM",
        "fast clube": "Fast-AM",
        "fast am": "Fast-AM",
        "ferroviaria": "Ferroviária-SP",
        "ferroviaria sp": "Ferroviária-SP",
        "ferroviario": "Ferroviário-CE",
        "ferroviario ce": "Ferroviário-CE",
        "ferroviario ma": "Ferroviário-MA",
        "ferroviario pr": "Ferroviário-PR",
        "figueirense": "Figueirense",
        "flamengo": "Flamengo",
        "flamengo rj": "Flamengo",
        "flamengo pi": "Flamengo-PI",
        "fluminense": "Fluminense",
        "fluminense rj": "Fluminense",
        "fluminense ba": "Fluminense-BA",
        "fonseca rj": "Fonseca-RJ",
        "fortaleza": "Fortaleza",
        "francana": "Francana",
        "ge maringa": "GE Maringá",
        "galicia": "Galícia",
        "gama": "Gama",
        "goias": "Goias",
        "goiania": "Goiânia",
        "goytacaz": "Goytacaz",
        "gremio": "Gremio",
        "gremio prudente": "Gremio Prudente",
        "gremio maringa": "Grêmio Maringá",
        "guanabara df": "Guanabara-DF",
        "guarani": "Guarani",
        "guara": "Guará",
        "hercilio luz": "Hercílio Luz",
        "inter de limeira": "Inter de Limeira",
        "inter de santa maria": "Inter de Santa Maria",
        "internacional": "Internacional",
        "internacional rs": "Internacional",
        "inter": "Internacional",
        "internacional sc": "Internacional-SC",
        "internacional sp": "Inter de Limeira",
        "internacional sm": "Inter de Santa Maria",
        "ipatinga": "Ipatinga",
        "itabaiana": "Itabaiana",
        "itabuna": "Itabuna-BA",
        "itabuna ba": "Itabuna-BA",
        "itumbiara": "Itumbiara",
        "j malucelli": "J.Malucelli",
        "joinville": "Joinville",
        "juventude": "Juventude",
        "juventus sp": "Juventus-SP",
        "leonico": "Leônico-BA",
        "leonico ba": "Leônico-BA",
        "liga da marinha": "Liga da Marinha",
        "londrina": "Londrina",
        "manufatora": "Manufatora",
        "maranhao": "Maranhão",
        "metropol": "Metropol",
        "mirassol": "Mirassol",
        "mixto": "Mixto",
        "moto club": "Moto Club",
        "moto clube": "Moto Club",
        "nacional": "Nacional-AM",
        "nacional am": "Nacional-AM",
        "nautico": "Nautico",
        "noroeste": "Noroeste",
        "novo hamburgo": "Novo Hamburgo",
        "olaria": "Olaria",
        "olympico am": "Olímpico-AM",
        "olimpico": "Olímpico-AM",
        "operario cg": "Operário-MS",
        "operario cg mt": "Operário-MS",
        "operario ms": "Operário-MS",
        "operario mt": "Operário-MT",
        "operario pr": "Operário-PR",
        "palmeiras": "Palmeiras",
        "parana": "Parana",
        "paula ramos sc": "Paula Ramos-SC",
        "paysandu": "Paysandu",
        "perdigao": "Perdigão",
        "piaui": "Piauí",
        "pinheiros pr": "Pinheiros-PR",
        "ponte preta": "Ponte Preta",
        "portuguesa": "Portuguesa",
        "portuguesa sp": "Portuguesa",
        "potiguar": "Potiguar",
        "rabello df": "Rabello-DF",
        "remo": "Remo",
        "rio branco": "Rio Branco-ES",
        "rio branco es": "Rio Branco-ES",
        "rio branco rj": "Rio Branco-RJ",
        "rio negro": "Rio Negro-AM",
        "rio negro am": "Rio Negro-AM",
        "river": "River-PI",
        "river pi": "River-PI",
        "sampaio correa": "Sampaio Corrêa",
        "santa cruz": "Santa Cruz",
        "santa cruz pe": "Santa Cruz",
        "santa cruz se": "Santa Cruz-SE",
        "santo andre": "Santo Andre",
        "santo antonio es": "Santo Antônio-ES",
        "santos": "Santos",
        "sao caetano": "Sao Caetano",
        "sao jose sp": "São José-SP",
        "sao paulo": "Sao Paulo",
        "sao paulo sp": "Sao Paulo",
        "sao paulo rs": "São Paulo-RS",
        "sao bento": "São Bento",
        "sergipe": "Sergipe",
        "siderurgica": "Siderúrgica",
        "sobradinho": "Sobradinho",
        "sport": "Sport",
        "sport recife": "Sport",
        "taguatinga": "Taguatinga",
        "tiradentes": "Tiradentes-PI",
        "tiradentes pi": "Tiradentes-PI",
        "treze": "Treze",
        "tuna luso": "Tuna Luso",
        "uberaba": "Uberaba",
        "uberlandia": "Uberlândia",
        "uniao sao joao": "União São João",
        "vasco": "Vasco",
        "vasco da gama": "Vasco",
        "vasco da gama rj": "Vasco",
        "vila nova": "Vila Nova-GO",
        "vila nova go": "Vila Nova-GO",
        "villa nova mg": "Villa Nova-MG",
        "vitoria": "Vitoria",
        "vitoria es": "Vitória-ES",
        "volta redonda": "Volta Redonda",
        "xv de jau": "XV de Jaú",
        "xv novembro j": "XV de Jaú",
        "xv de novembro j": "XV de Jaú",
        "xv de piracicaba": "XV de Piracicaba",
        "xv novembro p": "XV de Piracicaba",
        "xv novembro p 18": "XV de Piracicaba",
        "xv de novembro p": "XV de Piracicaba",
        "agua verde": "Água Verde",
    }
    return aliases


def infer_uf(nome: str) -> str | None:
    if "-" in nome:
        suffix = nome.rsplit("-", 1)[-1]
        if len(suffix) == 2 and suffix.isalpha():
            return suffix.upper()
    known = {
        "ABC": "RN", "ASA": "AL", "Alecrim": "RN", "Bahia": "BA", "Bangu": "RJ",
        "Barueri": "SP", "Botafogo-RJ": "RJ", "Bragantino": "SP", "Corinthians": "SP",
        "Coritiba": "PR", "Cruzeiro": "MG", "Flamengo": "RJ", "Fluminense": "RJ",
        "Fortaleza": "CE", "Goias": "GO", "Gremio": "RS", "Guarani": "SP",
        "Internacional": "RS", "Juventude": "RS", "Nautico": "PE", "Palmeiras": "SP",
        "Paysandu": "PA", "Ponte Preta": "SP", "Portuguesa": "SP", "Santos": "SP",
        "Sao Paulo": "SP", "Sport": "PE", "Vasco": "RJ", "Vitoria": "BA",
    }
    return known.get(nome)


def read_classificacao() -> list[dict[str, object]]:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["classificacao"]
    headers = [c.value for c in ws[1]]
    rows = []
    aliases = canonical_seed()
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not values or values[0] is None:
            continue
        row = dict(zip(headers, values))
        origem = str(row["clube"])
        k = key(origem)
        canonical = aliases.get(k)
        if not canonical:
            row["clube_canonico"] = None
            row["alias_key"] = k
        else:
            row["clube_canonico"] = canonical
            row["alias_key"] = k
        rows.append(row)
    return rows


def write_report(rows: list[dict[str, object]], con: sqlite3.Connection) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "pendencias"
    ws.append(["clube_origem", "alias_key", "ocorrencias"])
    missing = {}
    for row in rows:
        if not row.get("clube_canonico"):
            missing[row["clube"]] = missing.get(row["clube"], 0) + 1
    for clube, qtd in sorted(missing.items()):
        ws.append([clube, key(clube), qtd])

    ws2 = wb.create_sheet("clubes_novos")
    ws2.append(["clube_canonico", "uf", "ocorrencias"])
    existing = {r[0] for r in con.execute("SELECT nome FROM dim_clube")}
    counts = {}
    for row in rows:
        canonical = row.get("clube_canonico")
        if canonical and canonical not in existing:
            counts[canonical] = counts.get(canonical, 0) + 1
    for clube, qtd in sorted(counts.items()):
        ws2.append([clube, infer_uf(clube), qtd])

    ws3 = wb.create_sheet("resumo")
    ws3.append(["metrica", "valor"])
    ws3.append(["linhas_excel", len(rows)])
    ws3.append(["clubes_origem_unicos", len({r["clube"] for r in rows})])
    ws3.append(["clubes_canonicos_unicos", len({r["clube_canonico"] for r in rows if r.get("clube_canonico")} )])
    ws3.append(["pendencias_alias", len(missing)])
    ws3.append(["clubes_novos_dim_clube", len(counts)])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for col in sheet.columns:
            letter = get_column_letter(col[0].column)
            sheet.column_dimensions[letter].width = min(max(len(str(c.value or "")) for c in col) + 2, 60)
        sheet.freeze_panes = "A2"
    wb.save(REPORT)


def ensure_schema(con: sqlite3.Connection) -> None:
    con.executescript(
        """
        PRAGMA foreign_keys = ON;

        CREATE TABLE IF NOT EXISTS dim_edicao_nacional (
            edicao_nacional_id INTEGER PRIMARY KEY AUTOINCREMENT,
            edicao_id          TEXT NOT NULL UNIQUE,
            temporada_id       INTEGER NOT NULL,
            competicao_nome    TEXT NOT NULL,
            tipo               TEXT,
            fonte_principal    TEXT,
            observacao         TEXT,
            num_clubes         INTEGER
        );

        CREATE TABLE IF NOT EXISTS dim_clube_alias (
            alias               TEXT PRIMARY KEY,
            alias_normalizado   TEXT NOT NULL,
            clube_id            INTEGER NOT NULL REFERENCES dim_clube(clube_id),
            nome_canonico       TEXT NOT NULL,
            fonte               TEXT
        );

        CREATE TABLE IF NOT EXISTS stg_classificacao_final_codex (
            temporada_id     INTEGER NOT NULL,
            edicao_id        TEXT NOT NULL,
            competicao_nome  TEXT NOT NULL,
            posicao          INTEGER NOT NULL,
            clube_origem     TEXT NOT NULL,
            clube_canonico   TEXT NOT NULL,
            clube_id         INTEGER NOT NULL,
            fonte            TEXT,
            observacao       TEXT
        );

        CREATE TABLE IF NOT EXISTS fato_classificacao_final_nacional (
            edicao_nacional_id INTEGER NOT NULL REFERENCES dim_edicao_nacional(edicao_nacional_id),
            temporada_id       INTEGER NOT NULL,
            clube_id           INTEGER NOT NULL REFERENCES dim_clube(clube_id),
            posicao            INTEGER NOT NULL,
            fonte              TEXT,
            observacao         TEXT,
            PRIMARY KEY (edicao_nacional_id, clube_id),
            UNIQUE (edicao_nacional_id, posicao)
        );

        CREATE INDEX IF NOT EXISTS idx_class_final_temp
            ON fato_classificacao_final_nacional(temporada_id, posicao);
        CREATE INDEX IF NOT EXISTS idx_class_final_clube
            ON fato_classificacao_final_nacional(clube_id);

        DROP VIEW IF EXISTS vw_classificacao_final_historica;
        CREATE VIEW vw_classificacao_final_historica AS
        SELECT
            e.temporada_id,
            e.edicao_id,
            e.competicao_nome,
            e.tipo,
            f.posicao,
            c.clube_id,
            c.nome AS clube,
            c.uf,
            f.fonte,
            f.observacao
        FROM fato_classificacao_final_nacional f
        JOIN dim_edicao_nacional e
          ON e.edicao_nacional_id = f.edicao_nacional_id
        JOIN dim_clube c
          ON c.clube_id = f.clube_id;
        """
    )


def upsert_clube(con: sqlite3.Connection, nome: str) -> int:
    row = con.execute("SELECT clube_id FROM dim_clube WHERE nome = ?", (nome,)).fetchone()
    if row:
        return int(row[0])
    con.execute("INSERT INTO dim_clube(nome, uf) VALUES (?, ?)", (nome, infer_uf(nome)))
    return int(con.execute("SELECT last_insert_rowid()").fetchone()[0])


def tipo_competicao(nome: str) -> str:
    if "Taça Brasil" in nome:
        return "taca_brasil"
    if "Roberto Gomes Pedrosa" in nome:
        return "robertao"
    if "Torneio dos Campeões" in nome:
        return "torneio_dos_campeoes"
    return "brasileirao"


def import_rows(rows: list[dict[str, object]], con: sqlite3.Connection) -> None:
    aliases = canonical_seed()
    con.execute("DELETE FROM fato_classificacao_final_nacional")
    con.execute("DELETE FROM stg_classificacao_final_codex")
    con.execute("DELETE FROM dim_edicao_nacional")
    con.execute("DELETE FROM dim_clube_alias")

    canonical_to_id: dict[str, int] = {}
    for row in rows:
        canonical = str(row["clube_canonico"])
        canonical_to_id[canonical] = upsert_clube(con, canonical)

    for alias_key, canonical in aliases.items():
        if canonical in canonical_to_id:
            con.execute(
                """
                INSERT OR REPLACE INTO dim_clube_alias
                    (alias, alias_normalizado, clube_id, nome_canonico, fonte)
                VALUES (?, ?, ?, ?, ?)
                """,
                (alias_key, alias_key, canonical_to_id[canonical], canonical, "classificacao_codex.xlsx"),
            )

    edicoes = {}
    for row in rows:
        edicao_id = str(row["edicao_id"])
        if edicao_id not in edicoes:
            edicoes[edicao_id] = {
                "temporada": int(row["temporada"]),
                "competicao": str(row["competicao"]),
                "fonte": str(row.get("fonte") or ""),
                "observacao": str(row.get("observacao") or ""),
                "num_clubes": 0,
            }
        edicoes[edicao_id]["num_clubes"] += 1

    for edicao_id, ed in edicoes.items():
        con.execute(
            """
            INSERT INTO dim_edicao_nacional
                (edicao_id, temporada_id, competicao_nome, tipo, fonte_principal, observacao, num_clubes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                edicao_id,
                ed["temporada"],
                ed["competicao"],
                tipo_competicao(str(ed["competicao"])),
                ed["fonte"],
                ed["observacao"],
                ed["num_clubes"],
            ),
        )

    edicao_db = {
        edicao_id: edicao_nacional_id
        for edicao_nacional_id, edicao_id in con.execute(
            "SELECT edicao_nacional_id, edicao_id FROM dim_edicao_nacional"
        )
    }

    for row in rows:
        canonical = str(row["clube_canonico"])
        clube_id = canonical_to_id[canonical]
        con.execute(
            """
            INSERT INTO stg_classificacao_final_codex
                (temporada_id, edicao_id, competicao_nome, posicao, clube_origem,
                 clube_canonico, clube_id, fonte, observacao)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(row["temporada"]),
                str(row["edicao_id"]),
                str(row["competicao"]),
                int(row["posicao"]),
                str(row["clube"]),
                canonical,
                clube_id,
                str(row.get("fonte") or ""),
                str(row.get("observacao") or ""),
            ),
        )
        con.execute(
            """
            INSERT INTO fato_classificacao_final_nacional
                (edicao_nacional_id, temporada_id, clube_id, posicao, fonte, observacao)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                edicao_db[str(row["edicao_id"])],
                int(row["temporada"]),
                clube_id,
                int(row["posicao"]),
                str(row.get("fonte") or ""),
                str(row.get("observacao") or ""),
            ),
        )


def validate(con: sqlite3.Connection) -> list[tuple[str, int]]:
    checks = []
    checks.append(("linhas_fato", con.execute("SELECT COUNT(*) FROM fato_classificacao_final_nacional").fetchone()[0]))
    checks.append(("edicoes", con.execute("SELECT COUNT(*) FROM dim_edicao_nacional").fetchone()[0]))
    checks.append(("clubes_dim", con.execute("SELECT COUNT(*) FROM dim_clube").fetchone()[0]))
    checks.append(("posicoes_duplicadas", con.execute(
        """
        SELECT COUNT(*) FROM (
          SELECT edicao_nacional_id, posicao, COUNT(*) qtd
          FROM fato_classificacao_final_nacional
          GROUP BY edicao_nacional_id, posicao
          HAVING qtd > 1
        )
        """
    ).fetchone()[0]))
    checks.append(("clubes_duplicados_na_edicao", con.execute(
        """
        SELECT COUNT(*) FROM (
          SELECT edicao_nacional_id, clube_id, COUNT(*) qtd
          FROM fato_classificacao_final_nacional
          GROUP BY edicao_nacional_id, clube_id
          HAVING qtd > 1
        )
        """
    ).fetchone()[0]))
    checks.append(("edicoes_sem_posicao_1", con.execute(
        """
        SELECT COUNT(*) FROM dim_edicao_nacional e
        WHERE NOT EXISTS (
          SELECT 1 FROM fato_classificacao_final_nacional f
          WHERE f.edicao_nacional_id = e.edicao_nacional_id AND f.posicao = 1
        )
        """
    ).fetchone()[0]))
    return [(name, int(value)) for name, value in checks]


def main() -> None:
    if not DB.exists():
        raise SystemExit(f"Banco nao encontrado: {DB}")
    if not XLSX.exists():
        raise SystemExit(f"Planilha nao encontrada: {XLSX}")

    rows = read_classificacao()
    pending = [r for r in rows if not r.get("clube_canonico")]

    con = sqlite3.connect(DB)
    con.execute("PRAGMA foreign_keys = ON")
    write_report(rows, con)
    if pending:
        con.close()
        raise SystemExit(
            f"Ha {len(pending)} linhas com clube sem alias. Revise {REPORT} antes de importar."
        )

    if not BACKUP.exists():
        shutil.copy2(DB, BACKUP)

    with con:
        ensure_schema(con)
        import_rows(rows, con)
        checks = validate(con)

    con.close()
    print(f"OK importado em {DB}")
    print(f"Backup: {BACKUP}")
    print(f"Relatorio: {REPORT}")
    for name, value in checks:
        print(f"{name}: {value}")
    print(f"Executado em: {datetime.now().isoformat(timespec='seconds')}")


if __name__ == "__main__":
    main()
