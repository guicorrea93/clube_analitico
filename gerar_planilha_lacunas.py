from __future__ import annotations

import csv
import sqlite3
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DB_PATH = Path("db/brasileirao.db")
OUTPUT_PATH = Path("docs/lacunas_para_pesquisa.xlsx")
ERROR_FILES = [
    Path("data/transfermarkt_eventos_2003_2006_erros.csv"),
    Path("data/transfermarkt_tecnicos_formacoes_2006_2013_erros.csv"),
    Path("data/espn_eventos_2006_2020_erros.csv"),
    Path("data/espn_cartoes_2025.csv"),
]


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, header in enumerate(headers, start=1):
        values = [str(header)] + [str(row.get(header, "")) for row in rows[:500]]
        width = min(max(len(v) for v in values) + 2, 55)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def load_matches(cur: sqlite3.Cursor) -> dict[int, dict]:
    rows = cur.execute(
        """
        SELECT p.partida_id, p.temporada_id, p.rodada, p.data,
               hm.nome AS mandante, aw.nome AS visitante,
               p.gols_mandante, p.gols_visitante,
               p.tecnico_mand_id, p.tecnico_vis_id,
               p.formacao_mandante, p.formacao_visitante
        FROM fato_partida p
        JOIN dim_clube hm ON hm.clube_id = p.mandante_id
        JOIN dim_clube aw ON aw.clube_id = p.visitante_id
        ORDER BY p.temporada_id, p.rodada, p.data, p.partida_id
        """
    ).fetchall()
    return {row["partida_id"]: dict(row) for row in rows}


def tecnico_formacao_rows(matches: dict[int, dict]) -> list[dict]:
    rows = []
    for match in matches.values():
        missing = []
        if match["tecnico_mand_id"] is None:
            missing.append("tecnico_mandante")
        if match["tecnico_vis_id"] is None:
            missing.append("tecnico_visitante")
        if not match["formacao_mandante"]:
            missing.append("formacao_mandante")
        if not match["formacao_visitante"]:
            missing.append("formacao_visitante")
        if missing:
            rows.append({
                "temporada": match["temporada_id"],
                "rodada": match["rodada"],
                "data": match["data"],
                "partida_id": match["partida_id"],
                "mandante": match["mandante"],
                "visitante": match["visitante"],
                "placar": f'{match["gols_mandante"]}-{match["gols_visitante"]}',
                "o_que_falta": ", ".join(missing),
                "onde_procurei": "Transfermarkt; ESPN quando aplicavel",
                "observacao": "Arrecadacao ignorada por solicitacao.",
            })
    return rows


def gols_rows(cur: sqlite3.Cursor, matches: dict[int, dict]) -> list[dict]:
    counts = dict(cur.execute("SELECT partida_id, COUNT(*) FROM fato_gol GROUP BY partida_id").fetchall())
    rows = []
    for match in matches.values():
        esperado = match["gols_mandante"] + match["gols_visitante"]
        atual = counts.get(match["partida_id"], 0)
        if atual != esperado:
            rows.append({
                "temporada": match["temporada_id"],
                "rodada": match["rodada"],
                "data": match["data"],
                "partida_id": match["partida_id"],
                "mandante": match["mandante"],
                "visitante": match["visitante"],
                "placar": f'{match["gols_mandante"]}-{match["gols_visitante"]}',
                "gols_esperados_pelo_placar": esperado,
                "gols_no_banco": atual,
                "diferenca": esperado - atual,
                "onde_procurei": "Transfermarkt para 2003-2005; ESPN para 2006-2013; base local para 2014+",
                "observacao": "Priorize completar os autores/minutos dos gols desta partida.",
            })
    return rows


def cartoes_detail_rows(cur: sqlite3.Cursor, matches: dict[int, dict]) -> list[dict]:
    rows = cur.execute(
        """
        SELECT partida_id,
               COUNT(*) AS cartoes,
               SUM(posicao_id IS NULL) AS sem_posicao,
               SUM(num_camisa IS NULL OR num_camisa = '') AS sem_camisa
        FROM fato_cartao
        GROUP BY partida_id
        HAVING sem_posicao > 0 OR sem_camisa > 0
        """
    ).fetchall()
    out = []
    for row in rows:
        match = matches[row["partida_id"]]
        out.append({
            "temporada": match["temporada_id"],
            "rodada": match["rodada"],
            "data": match["data"],
            "partida_id": match["partida_id"],
            "mandante": match["mandante"],
            "visitante": match["visitante"],
            "placar": f'{match["gols_mandante"]}-{match["gols_visitante"]}',
            "cartoes_no_banco": row["cartoes"],
            "cartoes_sem_posicao": row["sem_posicao"] or 0,
            "cartoes_sem_camisa": row["sem_camisa"] or 0,
            "onde_procurei": "Transfermarkt; ESPN",
            "observacao": "O cartao existe, mas faltam posicao e/ou numero de camisa.",
        })
    return sorted(out, key=lambda r: (r["temporada"], r["rodada"], r["partida_id"]))


def source_error_rows() -> list[dict]:
    rows = []
    for path in ERROR_FILES:
        if not path.exists() or path.name == "espn_cartoes_2025.csv":
            continue
        with path.open(encoding="utf-8", newline="") as fp:
            for item in csv.DictReader(fp):
                key = tuple(item.items())
                rows.append({
                    "arquivo_fonte": str(path),
                    "temporada": item.get("temporada", ""),
                    "rodada": item.get("rodada", ""),
                    "id_fonte": item.get("match_id") or item.get("source_event_id") or "",
                    "erro_ou_lacuna": item.get("erro", ""),
                    "o_que_fazer": "Conferir manualmente esta partida/evento em outra fonte.",
                })
    seen = set()
    unique = []
    for row in rows:
        key = tuple(row.items())
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique


def main() -> int:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    try:
        cur = con.cursor()
        matches = load_matches(cur)
        tecnicos = tecnico_formacao_rows(matches)
        gols = gols_rows(cur, matches)
        cartoes = cartoes_detail_rows(cur, matches)
        erros = source_error_rows()

        resumo_counts = Counter()
        resumo_counts["partidas com tecnico/formacao faltante"] = len(tecnicos)
        resumo_counts["partidas com gols divergentes"] = len(gols)
        resumo_counts["partidas com cartoes sem detalhe"] = len(cartoes)
        resumo_counts["erros/lacunas de fontes online"] = len(erros)

        resumo = [
            {"categoria": categoria, "quantidade": quantidade, "observacao": "Arrecadacao nao considerada nesta planilha."}
            for categoria, quantidade in resumo_counts.items()
        ]

        by_season = defaultdict(lambda: Counter())
        for row in tecnicos:
            by_season[row["temporada"]]["tecnico/formacao"] += 1
        for row in gols:
            by_season[row["temporada"]]["gols"] += 1
        for row in cartoes:
            by_season[row["temporada"]]["cartoes_sem_detalhe"] += 1
        resumo_temporada = [
            {
                "temporada": season,
                "partidas_tecnico_formacao": counts["tecnico/formacao"],
                "partidas_gols_divergentes": counts["gols"],
                "partidas_cartoes_sem_detalhe": counts["cartoes_sem_detalhe"],
            }
            for season, counts in sorted(by_season.items())
        ]

        wb = Workbook()
        wb.remove(wb.active)
        add_sheet(wb, "Resumo", ["categoria", "quantidade", "observacao"], resumo)
        add_sheet(wb, "Resumo por Temporada", ["temporada", "partidas_tecnico_formacao", "partidas_gols_divergentes", "partidas_cartoes_sem_detalhe"], resumo_temporada)
        add_sheet(wb, "Tecnicos Formacoes", ["temporada", "rodada", "data", "partida_id", "mandante", "visitante", "placar", "o_que_falta", "onde_procurei", "observacao"], tecnicos)
        add_sheet(wb, "Gols Divergentes", ["temporada", "rodada", "data", "partida_id", "mandante", "visitante", "placar", "gols_esperados_pelo_placar", "gols_no_banco", "diferenca", "onde_procurei", "observacao"], gols)
        add_sheet(wb, "Cartoes Sem Detalhe", ["temporada", "rodada", "data", "partida_id", "mandante", "visitante", "placar", "cartoes_no_banco", "cartoes_sem_posicao", "cartoes_sem_camisa", "onde_procurei", "observacao"], cartoes)
        add_sheet(wb, "Erros Fontes Online", ["arquivo_fonte", "temporada", "rodada", "id_fonte", "erro_ou_lacuna", "o_que_fazer"], erros)
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        wb.save(OUTPUT_PATH)
        print(f"Planilha criada: {OUTPUT_PATH}")
        return 0
    finally:
        con.close()


if __name__ == "__main__":
    raise SystemExit(main())
